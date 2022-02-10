import os
import sys

import openpyxl
import requests

path = os.path.abspath(os.path.dirname(os.path.dirname(__file__)))
sys.path.append(path)


class HandExcel:
    # 加载excel
    def load_excel(self):
        open_exel = openpyxl.load_workbook(path + '/excel/test1.xlsx')
        return open_exel

    # 获取单元格内容
    def get_cell_value(self, sheet_name, row, cols):
        data = self.load_excel()[sheet_name].cell(row=row, column=cols).value
        return data

    # 获取某一行的内容
    def get_rows_value(self, sheet_name, row):
        row_list = []
        for i in self.load_excel()[sheet_name][row]:
            row_list.append(i.value)
        return row_list

    # 写入数据
    def write_excel(self, sheet_name, row, cols, value):
        ww = self.load_excel()
        ws = ww[sheet_name]
        ws.cell(row=row, column=cols).value = value
        ww.save(path + '/excel/test1.xlsx')


excel_data = HandExcel()

if __name__ == '__main__':
    data = excel_data.get_rows_value('登陆', 2)
    method = data[0]
    url = data[1]
    header1 = eval(data[2])
    cookie = data[3]
    params1 = '_=1644324708618'
    cookie_dict = {i.split("=")[0]: i.split("=")[-1] for i in cookie.split("; ")}
    print(method)
    print(type(url))
    print(type(header1))
    print(cookie)

    if method == 'get':
        re = requests.get(url, params=params1, headers=header1, cookies=cookie_dict, verify=False)
        print(re.json())
        excel_data.write_excel('登陆', 2, 5, str(re.json()))
    else:
        print('不匹配')